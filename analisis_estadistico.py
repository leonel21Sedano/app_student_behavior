import tkinter as tk
from tkinter import ttk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
from collections import Counter
import pandas as pd
import io
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as XLImage

class StatisticalAnalysisWindow:
    def __init__(self, parent, detection_data=None):
        self.parent = parent
        self.detection_data = detection_data
        self.colors = {
            'bg_primary': '#1a1a1a',
            'bg_secondary': '#2d2d2d',
            'bg_tertiary': '#3d3d3d',
            'accent': '#00d4aa',
            'text_primary': '#ffffff',
            'text_secondary': '#b2b2b2'
        }
        self.create_window()

    def create_window(self):
        self.window = tk.Toplevel(self.parent)
        self.window.title("📊 Análisis Estadístico de Comportamientos")
        self.window.geometry("1400x900")
        self.window.configure(bg=self.colors['bg_primary'])
        self.window.minsize(1000, 700)

        header_frame = tk.Frame(self.window, bg=self.colors['bg_secondary'], height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)

        title_label = tk.Label(header_frame,
                              text="📊 Análisis Estadístico de Comportamientos",
                              font=("Segoe UI", 20, "bold"),
                              bg=self.colors['bg_secondary'],
                              fg=self.colors['text_primary'])
        title_label.pack(pady=20)

        main_frame = tk.Frame(self.window, bg=self.colors['bg_primary'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        control_frame = tk.Frame(main_frame, bg=self.colors['bg_secondary'], height=60)
        control_frame.pack(fill=tk.X, pady=(0, 20))
        control_frame.pack_propagate(False)

        btn_frame = tk.Frame(control_frame, bg=self.colors['bg_secondary'])
        btn_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        btn_generate = tk.Button(btn_frame,
                                 text="🔄 Generar Análisis",
                                 command=self.generate_analysis,
                                 bg=self.colors['accent'],
                                 fg=self.colors['text_primary'],
                                 font=("Segoe UI", 11, "bold"),
                                 relief=tk.FLAT,
                                 padx=20, pady=8)
        btn_generate.pack(side=tk.LEFT, padx=(0, 10))

        btn_export = tk.Button(btn_frame,
                               text="💾 Exportar Gráficos",
                               command=self.export_graphs,
                               bg="#e67e22",
                               fg=self.colors['text_primary'],
                               font=("Segoe UI", 11, "bold"),
                               relief=tk.FLAT,
                               padx=20, pady=8)
        btn_export.pack(side=tk.LEFT)

        btn_excel = tk.Button(btn_frame,
                              text="📄 Guardar en Excel",
                              command=self.export_to_excel,
                              bg="#2ecc71",
                              fg=self.colors['text_primary'],
                              font=("Segoe UI", 11, "bold"),
                              relief=tk.FLAT,
                              padx=20, pady=8)
        btn_excel.pack(side=tk.LEFT, padx=(10,0))

        self.graphics_frame = tk.Frame(main_frame, bg=self.colors['bg_secondary'])
        self.graphics_frame.pack(fill=tk.BOTH, expand=True)

        self.initial_message = tk.Label(self.graphics_frame,
                                       text="📈\n\nGenera análisis para visualizar gráficos estadísticos",
                                       font=("Segoe UI", 16),
                                       bg=self.colors['bg_secondary'],
                                       fg=self.colors['text_secondary'])
        self.initial_message.pack(fill=tk.BOTH, expand=True)

    def generate_sample_data(self):
        if self.detection_data and 'detections' in self.detection_data:
            try:
                detections = self.detection_data['detections']
                classes = self.detection_data.get('classes', {})
                if detections is not None and getattr(detections, "size", 0) > 0:
                    class_ids = detections[:, 5].astype(int)
                    class_counts = Counter(class_ids)
                    real_data = {}
                    for class_id, count in class_counts.items():
                        class_name = classes.get(class_id, f"Class_{class_id}")
                        real_data[class_name] = count
                    if real_data:
                        return real_data
            except Exception:
                pass
        behaviors = ['Handraise', 'Read', 'Write', 'Stand', 'Talk', 'Turn', 'Bow', 'Head', 'Discuss']
        freqs = [37, 15, 82, 91, 27, 35, 36, 63, 33]
        return dict(zip(behaviors, freqs))

    def generate_analysis(self):
        for widget in self.graphics_frame.winfo_children():
            widget.destroy()
        try:
            data = self.generate_sample_data()
            if not data:
                self.show_error_message("No hay datos suficientes para generar el análisis")
                return
            plt.style.use('dark_background')
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
            fig.patch.set_facecolor('#2d2d2d')

            behaviors = list(data.keys())
            freqs = list(data.values())

            colors_bar = ['#3498db', '#1abc9c', '#f39c12', '#e74c3c', '#9b59b6',
                          '#2ecc71', '#f1c40f', '#95a5a6', '#34495e']

            bars = ax1.bar(behaviors, freqs, color=colors_bar)
            ax1.set_title('Frecuencia de Comportamientos (Gráfico de Barras)',
                          fontsize=14, fontweight='bold', color='white', pad=20)
            ax1.set_xlabel('Comportamientos', fontsize=12, color='white')
            ax1.set_ylabel('Frecuencia', fontsize=12, color='white')
            ax1.tick_params(axis='x', rotation=45, colors='white')
            ax1.tick_params(axis='y', colors='white')
            ax1.grid(True, alpha=0.3)

            for bar in bars:
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height + 1,
                         f'{int(height)}', ha='center', va='bottom',
                         fontweight='bold', color='white')

            ax2.pie(freqs, labels=behaviors, colors=colors_bar,
                    autopct='%1.1f%%', startangle=90)
            ax2.set_title('Distribución de Comportamientos (Gráfico de Pastel)',
                          fontsize=14, fontweight='bold', color='white', pad=20)

            plt.tight_layout()

            canvas = FigureCanvasTkAgg(fig, self.graphics_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

            self.create_stats_panel(data)

        except Exception as e:
            self.show_error_message(f"Error al generar análisis: {str(e)}")

    def show_error_message(self, message):
        error_label = tk.Label(self.graphics_frame,
                              text=f"⚠️\n\n{message}",
                              font=("Segoe UI", 14),
                              bg=self.colors['bg_secondary'],
                              fg="#e74c3c")
        error_label.pack(fill=tk.BOTH, expand=True)

    def create_stats_panel(self, data):
        stats_frame = tk.Frame(self.graphics_frame, bg=self.colors['bg_tertiary'], height=120)
        stats_frame.pack(fill=tk.X, padx=20, pady=10)
        stats_frame.pack_propagate(False)

        stats_title = tk.Label(stats_frame,
                              text="📈 Estadísticas Resumidas",
                              font=("Segoe UI", 14, "bold"),
                              bg=self.colors['bg_tertiary'],
                              fg=self.colors['text_primary'])
        stats_title.pack(pady=10)

        stats_content = tk.Frame(stats_frame, bg=self.colors['bg_tertiary'])
        stats_content.pack(fill=tk.X, padx=20)

        values = list(data.values())
        total = sum(values)
        mean = np.mean(values) if values else 0
        most_freq = max(data.items(), key=lambda x: x[1])
        least_freq = min(data.items(), key=lambda x: x[1])

        stats_text = f"Total de detecciones: {total} | Promedio: {mean:.1f} | Más frecuente: {most_freq[0]} ({most_freq[1]}) | Menos frecuente: {least_freq[0]} ({least_freq[1]})"

        stats_label = tk.Label(stats_content,
                              text=stats_text,
                              font=("Segoe UI", 11),
                              bg=self.colors['bg_tertiary'],
                              fg=self.colors['text_secondary'],
                              wraplength=1000)
        stats_label.pack()

    def export_graphs(self):
        from tkinter import filedialog, messagebox

        filename = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG files", "*.png"), ("PDF files", "*.pdf"), ("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar gráficos"
        )

        if filename:
            try:
                data = self.generate_sample_data()

                plt.style.use('dark_background')
                fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 8))
                fig.patch.set_facecolor('#2d2d2d')

                behaviors = list(data.keys())
                freqs = list(data.values())
                colors_bar = ['#3498db', '#1abc9c', '#f39c12', '#e74c3c', '#9b59b6',
                              '#2ecc71', '#f1c40f', '#95a5a6', '#34495e']

                bars = ax1.bar(behaviors, freqs, color=colors_bar)
                ax1.set_title('Frecuencia de Comportamientos', fontsize=16, fontweight='bold')
                ax1.set_xlabel('Comportamientos', fontsize=14)
                ax1.set_ylabel('Frecuencia', fontsize=14)
                ax1.tick_params(axis='x', rotation=45)

                for bar in bars:
                    height = bar.get_height()
                    ax1.text(bar.get_x() + bar.get_width()/2., height + 1,
                             f'{int(height)}', ha='center', va='bottom', fontweight='bold')

                ax2.pie(freqs, labels=behaviors, colors=colors_bar,
                        autopct='%1.1f%%', startangle=90)
                ax2.set_title('Distribución de Comportamientos', fontsize=16, fontweight='bold')

                plt.tight_layout()

                if filename.lower().endswith('.xlsx'):
                    df = pd.DataFrame({'Comportamiento': behaviors, 'Frecuencia': freqs})
                    df.to_excel(filename, index=False, sheet_name='Datos')

                    buf = io.BytesIO()
                    fig.savefig(buf, format='png', dpi=300, bbox_inches='tight', facecolor='#2d2d2d')
                    buf.seek(0)

                    wb = openpyxl.load_workbook(filename)
                    ws = wb.create_sheet('Grafico')

                    img = Image.open(buf)
                    img_bytes = io.BytesIO()
                    img.save(img_bytes, format='PNG')
                    img_bytes.seek(0)

                    xl_img = XLImage(img_bytes)
                    ws.add_image(xl_img, 'A1')
                    wb.save(filename)
                    plt.close(fig)
                    messagebox.showinfo("Éxito", f"Datos y gráfico exportados a Excel:\n{filename}")
                    return

                fig.savefig(filename, dpi=300, bbox_inches='tight', facecolor='#2d2d2d', edgecolor='none')
                plt.close(fig)
                messagebox.showinfo("Éxito", f"Gráficos exportados exitosamente a:\n{filename}")

            except Exception as e:
                messagebox.showerror("Error", f"Error al exportar: {str(e)}")

    def export_to_excel(self):
        from tkinter import filedialog, messagebox

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar datos y gráfico en Excel"
        )

        if not filename:
            return

        try:
            data = self.generate_sample_data()
            behaviors = list(data.keys())
            freqs = list(data.values())

            df = pd.DataFrame({'Comportamiento': behaviors, 'Frecuencia': freqs})
            df.to_excel(filename, index=False, sheet_name='Datos')

            plt.style.use('dark_background')
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 8))
            fig.patch.set_facecolor('#2d2d2d')

            colors_bar = ['#3498db', '#1abc9c', '#f39c12', '#e74c3c', '#9b59b6',
                          '#2ecc71', '#f1c40f', '#95a5a6', '#34495e']

            bars = ax1.bar(behaviors, freqs, color=colors_bar)
            ax1.set_title('Frecuencia de Comportamientos', fontsize=16, fontweight='bold')
            ax1.set_xlabel('Comportamientos', fontsize=14)
            ax1.set_ylabel('Frecuencia', fontsize=14)
            ax1.tick_params(axis='x', rotation=45)

            for bar in bars:
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height + 1,
                         f'{int(height)}', ha='center', va='bottom', fontweight='bold')

            ax2.pie(freqs, labels=behaviors, colors=colors_bar,
                    autopct='%1.1f%%', startangle=90)
            ax2.set_title('Distribución de Comportamientos', fontsize=16, fontweight='bold')

            plt.tight_layout()

            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=300, bbox_inches='tight', facecolor='#2d2d2d')
            buf.seek(0)

            wb = openpyxl.load_workbook(filename)
            ws = wb.create_sheet('Grafico')

            img = Image.open(buf)
            img_bytes = io.BytesIO()
            img.save(img_bytes, format='PNG')
            img_bytes.seek(0)

            xl_img = XLImage(img_bytes)
            ws.add_image(xl_img, 'A1')
            wb.save(filename)
            plt.close(fig)

            messagebox.showinfo("Éxito", f"Datos y gráfico guardados en:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel: {e}")

def open_statistical_analysis(parent, detection_data=None):
    return StatisticalAnalysisWindow(parent, detection_data)

def abrir_analisis_estadistico(parent, datos_deteccion=None):
    return open_statistical_analysis(parent, datos_deteccion)
